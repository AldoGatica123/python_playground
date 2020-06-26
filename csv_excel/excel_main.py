import openpyxl
from openpyxl.chart import BarChart, Reference
import xlsxwriter


def get_sheet_names():
    workbook = openpyxl.load_workbook('res/my_excel.xlsx')
    print('Workbook sheets:')
    print(workbook.get_sheet_names())


def get_sheet():
    workbook = openpyxl.load_workbook('res/my_excel.xlsx')
    print('Current sheet:')
    print(workbook.get_sheet_by_name('People'))


def get_cell():
    workbook = openpyxl.load_workbook('res/my_excel.xlsx')
    sheet = workbook.get_sheet_by_name('People')
    print('Current sheet:')
    print(sheet)
    print('Cell value:')
    print(sheet['B2'].value)


def update_cell():
    workbook = openpyxl.load_workbook('res/my_excel.xlsx')
    sheet = workbook['People']
    print('Current sheet:')
    print(sheet)
    print('Cell value:')
    print(sheet['B2'].value)
    print('New cell value:')
    sheet['B2'].value = 'Aldo'
    print(sheet['B2'].value)
    workbook.save('res/my_excel.xlsx')


def create_excel_file():
    workbook = xlsxwriter.Workbook('res/excel_writer.xlsx')
    print('Creating excel_writer.xlsx')
    workbook.add_worksheet(name='new sheet')
    workbook.close()


def create_new_sheet():
    workbook = openpyxl.load_workbook('res/excel_writer.xlsx')
    print('Adding new sheet to excel_writer.xlsx')
    workbook.create_sheet('woop woop')
    workbook.save('res/excel_writer.xlsx')


def write_to_excel():
    workbook = xlsxwriter.Workbook('res/excel_writer.xlsx')
    print('Writing to excel_writer.xlsx')
    sheet = workbook.add_worksheet(name='expenses')
    expenses = (['Rent', 1000], ['Gas', 100], ['Food', 300], ['Gym', 50])
    row = 0
    col = 0
    for item, cost in expenses:
        sheet.write(row, col, item)
        sheet.write(row, col + 1, cost)
        row += 1
    sheet.write(row, 0, 'Total')
    sheet.write(row, 1, '=SUM(B1:B4)')
    workbook.close()


def write_to_excel_with_format():
    workbook = xlsxwriter.Workbook('res/excel_writer.xlsx')
    format = workbook.add_format()
    format.set_bold()
    format.set_color('green')
    print('Writing with format to excel_writer.xlsx')
    sheet = workbook.add_worksheet(name='expenses')
    expenses = (['Rent', 1000], ['Gas', 100], ['Food', 300], ['Gym', 50])
    row = 0
    col = 0
    for item, cost in expenses:
        sheet.write(row, col, item)
        sheet.write(row, col + 1, cost)
        row += 1
    sheet.write(row, 0, 'Total')
    sheet.write(row, 1, '=SUM(B1:B4)')
    sheet.conditional_format('B1:KB5',
                             {'type': 'cell', 'criteria': '>=', 'value': 150, 'format': format})
    workbook.close()


def using_formulas():
    workbook = xlsxwriter.Workbook('res/excel_writer.xlsx')
    print('Using formulas in excel_writer.xlsx')
    sheet = workbook.add_worksheet(name='formulas')
    sheet.write_formula('A1', '=SUM(1,2,3)')
    workbook.close()


def make_line_chart():
    workbook = xlsxwriter.Workbook('res/excel_writer.xlsx')
    print('Adding charts in excel_writer.xlsx')
    sheet = workbook.add_worksheet()
    data = [10, 40, 50, 20, 10, 50]
    sheet.write_column('A1', data)
    chart = workbook.add_chart({'type': 'line'})
    chart.add_series({'values': 'Sheet1!$A$1:$A$6'})
    sheet.insert_chart('C1', chart)
    workbook.close()


def make_bar_chart():
    workbook = openpyxl.Workbook('res/excel_writer.xlsx')
    print('Adding charts in excel_writer.xlsx')
    sheet = workbook.create_sheet()
    data = [('Rent', 1000), ('Gas', 100), ('Food', 300), ('Gym', 50)]
    for row in data:
        sheet.append(row)
    chart = BarChart()
    chart.title = 'Expenses'
    chart.y_axis.title = 'Dollars'
    data = Reference(sheet, min_row=1, max_row=4, min_col=1, max_col=2)
    chart.add_data(data, from_rows=True, titles_from_data=True)
    sheet.add_chart(chart, 'D1')
    workbook.save('res/excel_writer.xlsx')


make_bar_chart()
